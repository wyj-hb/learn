
import xlrd
import numpy as np
import openpyxl as op
#问题1函数--------------------------

#该函数服务于问题1的权值计算，根据数据不同的材料种类返回对应的权值
#data表示要操纵的数据，i供应商编号
#根据计算可得A类的权值为0.4752，B类的权值为0.432，C类的权值为0.396
def panduan1(data,i):
    if(data[i][1] == 'A'):
        return 0.4752
    elif(data[i][1] == 'B'):
        return 0.4320
    else:
        return 0.3960
    pass
def channengbiao(filename,data,str):
    bg = op.load_workbook(filename)  # 应先将excel文件放入到工作目录下
    sheet = bg[str]  # “Sheet1”表示将数据写入到excel文件的sheet1下
    for i in range(len(data)):#403
        for j in range(len(data[i])):#242
            if (i == 0):
                sheet.cell(i + 1, j + 1, data[i][j])
                continue
            if (j == 0 or j == 1):
                sheet.cell(i + 1, j + 1, data[i][j])
                continue
            a = panduan1(data,i)
            sheet.cell(i+1,j+1,data[i][j]*a)
            pass
        pass
    bg.save(filename)  # 对文件进行保存
    pass
def wushijia(filename,data,str,data2):#data2中存放的是索引位置
    bg = op.load_workbook(filename)  # 应先将excel文件放入到工作目录下
    sheet = bg[str]  # “Sheet1”表示将数据写入到excel文件的sheet1下
    for i in range(51):  #50个
        for j in range(len(data[i])):  # 242
            if (i == 0):
                sheet.cell(i + 1, j + 1, data[i][j])
                continue
            bianhao = int(data2[i - 1][1]);
            if (j == 0 or j == 1):
                sheet.cell(i + 1, j + 1, data[bianhao][j])
                continue
            sheet.cell(i + 1, j + 1, data[bianhao][j])
            pass
        pass
    bg.save(filename)  # 对文件进行保存
    pass
#问题2 ---------------------

def wenti2shaixuan(filename,data1,len,data2,str,len1):#data1存的是编号 data2存的是完整的数据
    bg = op.load_workbook(filename)  # 应先将excel文件放入到工作目录下
    sheet = bg[str]  # “Sheet1”表示将数据写入到excel文件的sheet1下
    for i in range(len+1):
        for j in range(0,len1):
            if(i == 0):
                sheet.cell(i + 1, j + 1, data2[i][j])
                continue
            bianhao = int(data1[i-1][1]);
            if(j == 0 or j == 1):
                sheet.cell(i+1,j+1,data2[bianhao][j])
                continue
            sheet.cell(i+1, j+1, (data2[bianhao][j]))  # sheet.cell(1,1,num_list[0])表示将num_list列表的第0个数据1写入到excel表格的第一行第一列
    bg.save(filename)  # 对文件进行保存
    pass
def wenti2gonghuo(filename,data1):#data1存的是编号
    bg = op.load_workbook(filename)  # 应先将excel文件放入到工作目录下
    sheet = bg["Sheet1"]  # “Sheet1”表示将数据写入到excel文件的sheet1下
    for i in range(len(data1)):
        for j in range(0,24):
            sheet.cell(i + 2, j + 3, (data1[i][j]))
    bg.save(filename)  # 对文件进行保存
    pass
#问题三
def wenti3shaixuan(filename,data1,len,data2,str,len1):#data1存的是编号 data2存的是完整的数据
    bg = op.load_workbook(filename)  # 应先将excel文件放入到工作目录下
    sheet = bg[str]  # “Sheet1”表示将数据写入到excel文件的sheet1下
    for i in range(len+1):
        for j in range(0,len1):
            if(i == 0):
                sheet.cell(i + 1, j + 1, data2[i][j])
                continue
            bianhao = int(data1[i-1]);
            if(j == 0 or j == 1):
                sheet.cell(i+1,j+1,data2[bianhao][j])
                continue
            sheet.cell(i+1, j+1, (data2[bianhao][j]))  # sheet.cell(1,1,num_list[0])表示将num_list列表的第0个数据1写入到excel表格的第一行第一列
    bg.save(filename)  # 对文件进行保存
    pass
def wenti4shaixuan(filename,len,data2,str,len1):#data1存的是编号 data2存的是完整的数据
    bg = op.load_workbook(filename)  # 应先将excel文件放入到工作目录下
    sheet = bg[str]  # “Sheet1”表示将数据写入到excel文件的sheet1下
    for i in range(len+1):
        for j in range(0,len1):
            if(i == 0):
                sheet.cell(i + 1, j + 1, data2[i][j])
                continue
            if(j == 0 or j == 1):
                sheet.cell(i+1,j+1,data2[i][j])
                continue
            sheet.cell(i+1, j+1, (data2[i][j]))  # sheet.cell(1,1,num_list[0])表示将num_list列表的第0个数据1写入到excel表格的第一行第一列
    bg.save(filename)  # 对文件进行保存
    pass
#都用的
def shujudaoru(filename,data1,data2,len,str,len1,len3):#data1存放索引位置,
    bg = op.load_workbook(filename)  # 应先将excel文件放入到工作目录下
    num = 0
    sheet = bg[str]  # 表示将数据写入到excel文件的str表
    for i in range(len):#242
        for j in range(0,len1):#24
            k = int(data2[num][1])
            if(data1[i+1][j+2] == 0):
                continue
            sheet.cell(k + 6, j + 2, (data1[i+1][j+2]))
            pass
        num += 1
        if num == len3:
            break;
    bg.save(filename)  # 对文件进行保存
    pass
def shujudaoru1(filename,data1,data2,len,str,len1,len3):#data1存放索引位置,
    bg = op.load_workbook(filename)  # 应先将excel文件放入到工作目录下
    num = 0
    sheet = bg[str]  # 表示将数据写入到excel文件的str表
    for i in range(len):#242
        for j in range(0,len1):#24
            k = int(data2[num])
            if (data1[i + 1][j + 2] == 0):
                continue
            sheet.cell(k + 6, j + 2, (data1[i+1][j+2]))
            pass
        num += 1
        if num == len3:
            break;
    bg.save(filename)  # 对文件进行保存
    pass
def shujudaoru3(filename,data1,len,str,len1,len3):
    bg = op.load_workbook(filename)  # 应先将excel文件放入到工作目录下
    num = 0
    sheet = bg[str]  # 表示将数据写入到excel文件的str表
    for i in range(len):#242
        for j in range(0,len1):#24
            if (data1[i + 1][j + 2] == 0):
                continue
            sheet.cell(i + 7, j + 2, (data1[i+1][j+2]))
            pass
        num += 1
        if num == len3:
            break;
    bg.save(filename)  # 对文件进行保存
    pass
# data5,list,list8
def shurudaoru2(filename,data1,data2,data3,str,len1,x):#data1中存放的是数据，data2中存放的是优先选择的转运商,data3存的是索引位置
    bg = op.load_workbook(filename)  # 应先将excel文件放入到工作目录下
    num = 0
    sheet = bg[str]  # 表示将数据写入到excel文件的str表
    for i in range(24):#24周
        # sum = 6000;
        # bianhao = 0#编号0，从损耗率最低的转运商选起
        # for j in range(8):#每周的8个转运商
        sum = 6000;
        bianhao = 0  # 编号0，从损耗率最低的转运商选起
            #选择我们所选的36个供应商进行比较
        for k in range(len1):
            if(data1[k+1][i+2]<sum):
                #为空不写0
                if(data1[k+1][i+2] == 0):
                       continue
                if(x == 2):
                    hangshu = data3[k][1]  # 229
                if(x == 3):
                    hangshu = int(data3[k])  # 229
                sum = sum -data1[k+1][i+2]
                xiabiao = data2[bianhao][1]+i*8#3
                sheet.cell(hangshu+6, xiabiao + 1, (data1[k+1][i+2]))
            elif(data1[k+1][i+2] == sum):
                if (x == 2):
                    hangshu = data3[k][1]  # 229
                if (x == 3):
                    hangshu = int(data3[k])  # 229
                sum = 6000
                xiabiao = data2[bianhao][1]+i*8  # 3
                sheet.cell(hangshu + 6, xiabiao + 1, (data1[k+1][i+2]))
                bianhao+=1
            elif(data1[k+1][i+2] > sum):
                if (x == 2):
                    hangshu = data3[k][1]  # 229
                if (x == 3):
                    hangshu = int(data3[k])  # 229
                a = data1[k+1][i+2] - sum
                xiabiao = data2[bianhao][1]+i*8  # 3
                sheet.cell(hangshu + 6, xiabiao + 1, sum)
                bianhao += 1
                sum = 6000
                sum = sum - a
                if(bianhao == 8):
                    break
                xiabiao = data2[bianhao][1]+i*8 # 3
                sheet.cell(hangshu + 6, xiabiao + 1, a)
            pass
        pass
    bg.save(filename)  # 对文件进行保存
    pass
def shurudaoru4(filename,data1,data2,str,len1):#data1中存放的是数据，data2中存放的是优先选择的转运商
    bg = op.load_workbook(filename)  # 应先将excel文件放入到工作目录下
    num = 0
    sheet = bg[str]  # 表示将数据写入到excel文件的str表
    for i in range(24):#24周
        # sum = 6000;
        # bianhao = 0#编号0，从损耗率最低的转运商选起
        # for j in range(8):#每周的8个转运商
        sum = 6000;
        bianhao = 0  # 编号0，从损耗率最低的转运商选起
            #选择我们所选的36个供应商进行比较
        for k in range(len1):#402个商家
            if(data1[k+1][i+2]<sum):
                #为空不写0
                if(data1[k+1][i+2] == 0):
                       continue
                sum = sum -data1[k+1][i+2]
                xiabiao = data2[bianhao][1]+i*8#3
                sheet.cell(k+7, xiabiao + 1, (data1[k+1][i+2]))
            elif(data1[k+1][i+2] == sum):
                sum = 6000
                xiabiao = data2[bianhao][1]+i*8  # 3
                sheet.cell(k + 7, xiabiao + 1, (data1[k+1][i+2]))
                bianhao+=1
            elif(data1[k+1][i+2] > sum):
                a = data1[k+1][i+2] - sum
                xiabiao = data2[bianhao][1]+i*8  # 3
                sheet.cell(i + 7, xiabiao + 1, sum)
                bianhao += 1
                sum = 6000
                sum = sum - a
                xiabiao = data2[bianhao][1]+i*8 # 3
                sheet.cell(k + 7, xiabiao + 1, a)
            pass
        pass
    bg.save(filename)  # 对文件进行保存
    pass